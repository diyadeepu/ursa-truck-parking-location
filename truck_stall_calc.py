import pandas as pd
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SQ_FT_PER_STALL = int(input("Enter square footage per truck parking stall (e.g. 300): ").strip())

url = input("Paste your public Google Sheets link: ").strip()

try:
    SHEET_ID = url.split("/d/")[1].split("/")[0]
except IndexError:
    print("ERROR: Could not find sheet ID.")
    exit(1)

# Extract gid (tab ID) from URL if present
GID = "0"
if "gid=" in url:
    GID = url.split("gid=")[-1].split("#")[0].split("&")[0]

CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}"

df = pd.read_csv(CSV_URL)
df.columns = [str(c).strip() for c in df.columns]

print("Columns found:", list(df.columns))
print(f"Total rows loaded: {len(df)}")

# Flexible square footage column detection
def find_sq_ft_col(columns):
    for c in columns:
        norm = c.lower().replace(".", "").replace(" ", "").replace("_", "")
        for kw in ["squarefootage", "sqfootage", "squareft", "sqft", "squarefeet", "sqfeet", "footage"]:
            if kw in norm:
                return c
    return None

sq_ft_col = find_sq_ft_col(df.columns)

if sq_ft_col is None:
    print("\nERROR: Could not auto-detect square footage column.")
    print("Available columns:")
    for c in df.columns:
        print(f"  - '{c}'")
    manual = input("\nType the exact column name to use: ").strip()
    if manual in df.columns:
        sq_ft_col = manual
    else:
        print("Column not found. Exiting.")
        exit(1)

print(f"\nUsing square footage column: '{sq_ft_col}'")

# Clean and convert — handle commas, blanks, non-numeric
df[sq_ft_col] = pd.to_numeric(
    df[sq_ft_col].astype(str).str.replace(",", "").str.strip(),
    errors="coerce"
)

# Keep ALL rows; blank/zero sq ft = "N/A" stalls
def calc_stalls(val):
    if pd.isna(val) or val <= 0:
        return "N/A"
    return math.floor(val / SQ_FT_PER_STALL)

df["Truck Parking Stalls"] = df[sq_ft_col].apply(calc_stalls)

# Insert stalls column right after sq_ft_col
cols = list(df.columns)
cols.remove("Truck Parking Stalls")
cols.insert(cols.index(sq_ft_col) + 1, "Truck Parking Stalls")
output_df = df[cols]

valid_count = (output_df["Truck Parking Stalls"] != "N/A").sum()
print(f"Rows processed: {len(output_df)}")
print(f"Rows with valid square footage: {valid_count}")

# CSV export
output_df.to_csv("truck_stalls.csv", index=False)
print("\nSaved: truck_stalls.csv")

# Excel export
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Truck Parking Stalls"

header_fill  = PatternFill("solid", fgColor="2F4F7F")
header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
sq_ft_fill   = PatternFill("solid", fgColor="D6E4F0")
stall_fill   = PatternFill("solid", fgColor="E2EFDA")
na_fill      = PatternFill("solid", fgColor="F2F2F2")
label_font   = Font(name="Arial", bold=True, size=10)
cell_font    = Font(name="Arial", size=10)
center       = Alignment(horizontal="center", vertical="center")
left         = Alignment(horizontal="left",   vertical="center")
thin         = Side(style="thin", color="CCCCCC")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

all_cols        = list(output_df.columns)
sq_ft_excel_col = all_cols.index(sq_ft_col) + 1
stall_excel_col = all_cols.index("Truck Parking Stalls") + 1

# Header row
for col_idx, col_name in enumerate(all_cols, start=1):
    c = ws.cell(row=1, column=col_idx, value=col_name)
    c.font      = header_font
    c.fill      = header_fill
    c.alignment = center
    c.border    = border

# Data rows
for row_idx, row in output_df.iterrows():
    excel_row = row_idx + 2
    for col_idx, col_name in enumerate(all_cols, start=1):
        val = row[col_name]
        if isinstance(val, float) and math.isnan(val):
            val = ""
        c = ws.cell(row=excel_row, column=col_idx, value=val)
        c.border = border
        c.font   = cell_font

        if col_idx == sq_ft_excel_col:
            c.fill      = sq_ft_fill
            c.font      = label_font
            c.alignment = center
            if isinstance(val, (int, float)):
                c.number_format = "#,##0"
        elif col_idx == stall_excel_col:
            if val == "N/A":
                c.fill = na_fill
                c.font = Font(name="Arial", size=10, color="999999", italic=True)
            else:
                c.fill          = stall_fill
                c.font          = label_font
                c.number_format = "#,##0"
            c.alignment = center
        else:
            c.alignment = left

# Column widths
for col_idx, col_name in enumerate(all_cols, start=1):
    ltr = get_column_letter(col_idx)
    if col_name == sq_ft_col:
        ws.column_dimensions[ltr].width = 18
    elif col_name == "Truck Parking Stalls":
        ws.column_dimensions[ltr].width = 22
    elif col_name in ("Latitude", "Longitude"):
        ws.column_dimensions[ltr].width = 16
    elif col_name == "Name":
        ws.column_dimensions[ltr].width = 34
    else:
        ws.column_dimensions[ltr].width = 14

# Row heights
ws.row_dimensions[1].height = 30
for i in range(len(output_df)):
    ws.row_dimensions[i + 2].height = 18

ws.freeze_panes = "B2"

# Footer note
note_col = len(all_cols) + 2
note = ws.cell(row=1, column=note_col,
               value=f"Note: 1 stall = {SQ_FT_PER_STALL:,} sq ft (rounded down). Blank sq ft shown as N/A.")
note.font = Font(name="Arial", italic=True, size=9, color="666666")

wb.save("truck_stalls.xlsx")
print("Saved: truck_stalls.xlsx")
print(f"\nDone! 1 stall = {SQ_FT_PER_STALL:,} sq ft, rounded down. Blanks shown as N/A.")
