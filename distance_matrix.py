import pandas as pd
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

url = input("Paste your public Google Sheets link: ").strip()
try:
    SHEET_ID = url.split("/d/")[1].split("/")[0]
except IndexError:
    print("ERROR: Could not find sheet ID. Make sure the URL looks like:")
    print("  https://docs.google.com/spreadsheets/d/YOUR_ID/...")
    exit(1)
CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv"

# get data from Google Sheets, convert latitude and longitude to numeric, and drop rows with invalid coordinates
df = pd.read_csv(CSV_URL, usecols=[0, 1, 2, 3])
df.columns = ["ID", "Name", "Latitude", "Longitude"]
df["Latitude"]  = pd.to_numeric(df["Latitude"],  errors="coerce")
df["Longitude"] = pd.to_numeric(df["Longitude"], errors="coerce")
df = df.dropna(subset=["Latitude", "Longitude"]).reset_index(drop=True)
print(f"Loaded {len(df)} locations.")

# get distance in miles between two latitude and longitude points using the Haversine formula
def haversine(lat1, lon1, lat2, lon2):
    R = 3958.8
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    a = math.sin((lat2-lat1)/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin((lon2-lon1)/2)**2
    return round(2 * R * math.asin(math.sqrt(a)), 2)

# build distance matrix by applying the Haversine function to each pair of locations
names = df["Name"].tolist()
ids   = df["ID"].tolist()
n     = len(names)
matrix = [[0.0]*n for _ in range(n)]
for i in range(n):
    for j in range(n):
        if i != j:
            matrix[i][j] = haversine(
                df.loc[i,"Latitude"], df.loc[i,"Longitude"],
                df.loc[j,"Latitude"], df.loc[j,"Longitude"])

# print the distance matrix using pandas, with 2 decimal places and no truncation
dist_df = pd.DataFrame(matrix, index=names, columns=names)
pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
print("\n" + "="*60)
print("  HAVERSINE DISTANCE MATRIX (miles)")
print("="*60)
print(dist_df.to_string(float_format=lambda x: f"{x:.2f}"))

# export distance matrix to CSV (has no formating here)
dist_df.to_csv("distance_matrix.csv")
print("\nSaved: distance_matrix.csv")

# export distance matrix to Excel with formating
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Distance Matrix"

header_fill   = PatternFill("solid", fgColor="2F4F7F")
header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
label_fill    = PatternFill("solid", fgColor="D6E4F0")
label_font    = Font(name="Arial", bold=True, size=10)
cell_font     = Font(name="Arial", size=10)
diag_fill     = PatternFill("solid", fgColor="EEEEEE")
center        = Alignment(horizontal="center", vertical="center")
left          = Alignment(horizontal="left",   vertical="center")
thin          = Side(style="thin", color="CCCCCC")
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

# Corner cells for No., ID, Name columns
for col, label in enumerate(["No.", "ID", "Truck Parking Location"], start=1):
    c = ws.cell(row=1, column=col, value=label)
    c.font = header_font; c.fill = header_fill
    c.alignment = center; c.border = border

# Column headers (location names across row 1, starting at col 4)
for j, name in enumerate(names):
    cell = ws.cell(row=1, column=j+4, value=name)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border

# Row labels + data
for i in range(n):
    # No.
    no = ws.cell(row=i+2, column=1, value=i+1)
    no.font = label_font; no.fill = label_fill
    no.alignment = center; no.border = border
    # ID
    id_cell = ws.cell(row=i+2, column=2, value=int(ids[i]))
    id_cell.font = label_font; id_cell.fill = label_fill
    id_cell.alignment = center; id_cell.border = border
    # Name
    lbl = ws.cell(row=i+2, column=3, value=names[i])
    lbl.font = label_font; lbl.fill = label_fill
    lbl.alignment = left; lbl.border = border

    # Data cells
    for j in range(n):
        c = ws.cell(row=i+2, column=j+4)
        c.border = border; c.alignment = center; c.font = cell_font
        if i == j:
            c.value = "—"; c.fill = diag_fill
        else:
            c.value = matrix[i][j]; c.number_format = "0.00"

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 30
for j in range(n):
    ws.column_dimensions[get_column_letter(j+4)].width = 14

# Row heights
ws.row_dimensions[1].height = 60
for i in range(n):
    ws.row_dimensions[i+2].height = 18

# Freeze panes after No., ID, Name columns
ws.freeze_panes = "D2"

wb.save("distance_matrix.xlsx")
print("Saved: distance_matrix.xlsx\n")
print("Done! Both files are in the same folder as your script.")
